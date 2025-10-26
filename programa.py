import datetime as dt
import sys
import json
import csv
import sqlite3
from sqlite3 import Error
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

HOY = dt.datetime.today().date()
FECHA_MINIMA = HOY + dt.timedelta(days=2)
TURNOS = ("Matutino", "Vespertino", "Nocturno")
DB = "Reservaciones.db"

class ReservacionesDB:
    def __init__(self,db=DB):
        self.db = db

    def execute_select(self, query, params=()):
        try:
            with sqlite3.connect(self.db) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute(query,params)
                return cursor.fetchall() 
        except Error as e:
            print(f'Error de tipo: {e}')

    def execute_insert(self, query, params=()):
        try:
            with sqlite3.connect(self.db) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute(query,params)
                return cursor.lastrowid
        except Error as e:
            print(f'Error de tipo: {e}')

    def execute_update(self, query, params=()):
        try:
            with sqlite3.connect(self.db) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute(query,params)
        except Error as e:
            print(f'Error de tipo: {e}')

    def conexion(self,query):
        try:
            with sqlite3.connect(self.db) as conn:
                cursor = conn.cursor()
                cursor.execute(query)
        except Error as e:
            print(f'Error de tipo: {e}')

db = ReservacionesDB()

def conexion():
    db.conexion("CREATE TABLE IF NOT EXISTS clientes"
                    "(id_cliente INTEGER PRIMARY KEY," \
                    "nombre TEXT NOT NULL," \
                    "apellido TEXT NOT NULL);")
    db.conexion("CREATE TABLE IF NOT EXISTS salas" \
                    "(id_sala INTEGER PRIMARY KEY," \
                    "nombre TEXT NOT NULL," \
                    "cupo INTEGER NOT NULL CHECK (cupo > 0));")
    db.conexion("CREATE TABLE IF NOT EXISTS reservas" \
                    "(folio INTEGER PRIMARY KEY," \
                    "id_cliente INTEGER NOT NULL," \
                    "id_sala INTEGER NOT NULL," \
                    "fecha TEXT NOT NULL," \
                    "nombre_evento TEXT NOT NULL," \
                    "turno TEXT NOT NULL CHECK (turno IN ('Matutino','Vespertino','Nocturno'))," \
                    "UNIQUE (id_sala, fecha, turno)," \
                    "FOREIGN KEY (id_cliente) REFERENCES clientes(id_cliente)," \
                    "FOREIGN KEY (id_sala)    REFERENCES salas(id_sala));")


def titulo(nombre):
    print("\n" + "="*70)
    texto = nombre.center(66)
    print("**" + texto + "**")
    print("="*70) 

def titulo2(nombre):
    print("\n" + "="*90)
    texto = nombre.center(86)
    print("**" + texto + "**")
    print("="*90)

def sep(cantidad):
    print("-"*cantidad)

def sep2(cantidad):
    print("#"*cantidad)

def salto():
    print("\n")

def formato(fecha):
    return dt.datetime.strftime(fecha,"%m-%d-%Y")

def formato_iso(fecha):
    return dt.datetime.strftime(fecha, "%Y-%m-%d")

def convertir_iso(fecha):
    return dt.datetime.strptime(fecha, "%Y-%m-%d").date()

def entrada(texto, c="C", tipo=str, hoy=None):
    while True:
        valor = input(f"{texto} o {c} para cancelar: ").strip().title()
        if valor == c:
            print("Operación cancelada... Volviendo al menú.\n")
            return None
        if not valor and hoy == "S":
            print(f"\nBuscando para hoy {formato(HOY)}... \n")
            return HOY
        else:
            if not valor:
                print("No puede quedar vacío.\n")
                continue
        if tipo is str:
            if valor.isdigit():
                print("No se permite escribir solo números.\n")
                continue
            return valor
        elif tipo is int:
            if not valor.isdigit():
                print("Debe ingresar un número entero.\n")
                valor = -1
            return int(valor)
        elif tipo is dt.date:
            try:
                return dt.datetime.strptime(valor, "%m-%d-%Y").date()
            except ValueError:
                print("La fecha proporcionada no es válida, favor de usar el formato correcto (MM-DD-YYYY)\n")
                continue


def hay_registros(tabla):
    datos = db.execute_select(f"SELECT COUNT(*) FROM {tabla}")
    return datos[0][0] > 0

def listar_clientes():
    datos = db.execute_select("SELECT id_cliente, apellido, nombre FROM clientes ORDER BY apellido, nombre", ())
    return datos

def existe_cliente(id_cliente):
    datos = db.execute_select("SELECT 1 FROM clientes WHERE id_cliente=?", (id_cliente,))
    return datos

def listar_salas():
    datos = db.execute_select("SELECT id_sala, nombre, cupo FROM salas ORDER BY id_sala")
    return datos

def turnos_ocupados(id_sala, fecha):
    datos = db.execute_select("SELECT turno FROM reservas WHERE id_sala=? AND fecha=?",(id_sala, fecha))
    turnos = [turno['turno'] for turno in datos]
    return turnos

def existe_sala(id_sala):
    datos = db.execute_select("SELECT id_sala FROM salas WHERE id_sala=?", (id_sala,))
    return datos

def insertar_reserva(id_cliente, id_sala, fecha, turno, nombre_evento):
    datos = db.execute_insert("INSERT INTO reservas (id_cliente, id_sala, fecha, turno, nombre_evento) VALUES (?,?,?,?,?)",
                      (id_cliente, id_sala, fecha, turno, nombre_evento))
    return datos

def datos_cliente(id_cliente):
    datos = db.execute_select("SELECT apellido || ' ' || nombre FROM clientes WHERE id_cliente=?", (id_cliente,))
    return datos[0][0]

def datos_sala(id_sala):
    datos = db.execute_select("SELECT nombre, cupo FROM salas WHERE id_sala=?", (id_sala,))
    return datos[0]

def reservas_en_rango(fecha_ini, fecha_fin):
    datos = db.execute_select("SELECT R.folio, " \
    "R.fecha, " \
    "R.nombre_evento, " \
    "C.apellido || ' ' || C.nombre AS cliente, " \
    "S.nombre AS sala, " \
    "R.turno " \
    "FROM reservas R " \
    "INNER JOIN clientes C ON C.id_cliente = R.id_cliente " \
    "INNER JOIN salas S ON S.id_sala = R.id_sala " \
    "WHERE R.fecha BETWEEN ? AND ? " \
    "ORDER BY R.fecha, R.folio", (fecha_ini, fecha_fin))
    return datos

def existe_folio(folio, fecha_inicio, fecha_fin):
    datos = db.execute_select("SELECT folio FROM reservas WHERE folio=? AND fecha BETWEEN ? AND ?"
                              , (folio, fecha_inicio, fecha_fin))
    return datos

def actualizar_nombre_evento(folio, nuevo_nombre):
    db.execute_update("UPDATE reservas SET nombre_evento=? WHERE folio=?", (nuevo_nombre, folio))

def insertar_cliente(valores):
    datos = db.execute_insert("INSERT INTO clientes (nombre, apellido) VALUES(?,?)", valores)
    return datos

def insertar_sala(valores):
    datos = db.execute_insert("INSERT INTO salas (nombre, cupo) VALUES(?,?)", valores)
    return datos

def reservas_por_fecha(fecha):
    datos = db.execute_select("SELECT R.folio, " \
    "C.apellido || ' ' || C.nombre AS cliente, " \
    "S.nombre AS sala, " \
    "R.nombre_evento, " \
    "S.cupo, " \
    "R.turno " \
    "FROM reservas R INNER JOIN clientes C ON C.id_cliente = R.id_cliente " \
    "INNER JOIN salas S  ON S.id_sala = R.id_sala " \
    "WHERE R.fecha = ? ORDER BY R.folio", (fecha,))
    return datos


def obtener_clientes():
    titulo("Clientes registrados")
    print(f"{'Clave':<8} {'Apellidos':<20} {'Nombres':<15}")
    sep(70)
    for r in listar_clientes():
        print(f"{r['id_cliente']:<8} {r['apellido']:<20} {r['nombre']:<15}")
    sep(70)

def obtener_salas(fecha):
    fecha_iso = formato_iso(fecha)
    titulo2(f"Salas disponibles para {formato(fecha)}")
    print(f"{'Clave':<8} {'Sala':<20} {'Cupo':<15} {'Turno disponible':<20}")
    sep(90)
    for sala in listar_salas():
        ocupados = turnos_ocupados(sala["id_sala"], fecha_iso)
        libres = [t for t in TURNOS if t not in ocupados]
        if not libres:
            print(f"{sala['id_sala']:<8} {sala['nombre']:<20} {'':<15} {'SALA OCUPADA':<20}")
        else:
            print(f"{sala['id_sala']:<8} {sala['nombre']:<20} {sala['cupo']:<15} {' | '.join(libres):<20}")
    sep(90)

def mostrar_reserva(folio, id_cliente, id_sala, fecha, nombre_evento, turno):
    cliente = datos_cliente(id_cliente)
    sala, cupo = datos_sala(id_sala)
    titulo("Reservación registrada exitosamente")
    print(f"Folio:   {folio}")
    print(f"Cliente: {cliente}")
    print(f"Sala:    {sala} | Cupo: {cupo}")
    print(f"Fecha:   {formato(fecha)}")
    print(f"Evento:  {nombre_evento}")
    print(f"Turno:   {turno}")

def agregar_reservacion_turno(id_cliente, id_sala, fecha, turno_elegido):
    nombre_evento = entrada("Ingrese el nombre del evento")
    if nombre_evento is None:
        return
    fecha_iso = formato_iso(fecha)
    folio = insertar_reserva(id_cliente, id_sala, fecha_iso, turno_elegido, nombre_evento)
    if not folio:
        print(f"El turno {turno_elegido} ya está ocupado en esa sala y fecha.")
        return
    mostrar_reserva(folio, id_cliente, id_sala, fecha, nombre_evento, turno_elegido)

def mostrar_reservacion_fechas(fecha_inicio, fecha_fin, reservas):
    titulo2(f"Reservaciones para {formato(fecha_inicio)} - {formato(fecha_fin)}")
    print(f"{'Folio':<8}{'Cliente':<19}{'Fecha':<15}{'Nombre del evento':<20}{'Sala':<20}{'Turno':<10}")
    sep(90)
    for reserva in reservas:
        fecha_formateada = formato(convertir_iso(reserva["fecha"]))
        print(f"{reserva['folio']:<8}{reserva['cliente']:<19}{fecha_formateada:<15}{reserva['nombre_evento']:<20}{reserva['sala']:<20}{reserva['turno']:<10}")
    sep(90)

def validacion_domingo(fecha):
    if fecha.weekday() != 6:
        return fecha
    while True:
        respuesta = entrada("No se puede hacer reservaciones en Domingo - ¿Desea pasar la fecha al siguiente lunes? (S/N)")
        if respuesta is None:
            return
        if respuesta == "S":
            return fecha + dt.timedelta(days=1)
        elif respuesta == "N":
            print("Volviendo a introducir una fecha nueva...\n")
            return respuesta
        else:
            print("Opción no válida (S/N)\n")
            continue


def exportar_json(reservas,fecha):
    datos = []
    for reserva in reservas:
        datos.append({"folio": reserva["folio"],
            "cliente": reserva["cliente"],
            "sala": reserva["sala"],
            "cupo": reserva["cupo"],
            "nombre_evento": reserva["nombre_evento"],
            "turno": reserva["turno"]})
    with open(f"Reporte-{formato(fecha)}.json", "w") as archivo:
        json.dump(datos, archivo, indent=2)

def exportar_csv(reservas,fecha):
    with open(f"Reporte-{formato(fecha)}.csv", "w", encoding="latin1", newline="") as archivo:
        grabador = csv.writer(archivo, delimiter=",")
        grabador.writerow(("Folio", "Nombre_Cliente", "Sala", "Cupo", "Nombre_Evento", "Turno"))
        for reserva in reservas:
            grabador.writerow([reserva['folio'],
                               reserva['cliente'],
                               reserva['sala'],
                               reserva['cupo'],
                               reserva['nombre_evento'],
                               reserva['turno']])
        
def exportar_excel(reservas,fecha):
    libro = openpyxl.Workbook()
    hoja = libro["Sheet"]
    hoja.title = "Reservaciones"

    negritas = Font(bold=True)
    centrado = Alignment(horizontal="center", vertical="center")
    borde_grueso = Border(bottom=Side(border_style="thick", color="00078C"))
    

    celdaA1 = hoja["A1"]
    celdaA1.value = "Reporte de Reservaciones"
    celdaA1.font = Font(bold=True, size=16)
    celdaA1.alignment = centrado

    for columna in range(1,7):
        celda = hoja.cell(row=1, column=columna)
        celda.border = borde_grueso

    encabezados = {1:"Folio", 2:"Cliente", 3:"Sala", 4:"Cupo", 5:"Evento", 6:"Turno"}
    for columna,texto in encabezados.items(): 
        celdaA2 = hoja.cell(row=2, column=columna, value=texto) 
        celdaA2.alignment = centrado
        celdaA2.font = negritas
        celdaA2.border = borde_grueso

    for fila,reserva in enumerate(reservas, start=3):
        hoja.cell(row=fila, column=1, value=reserva['folio']).alignment = centrado
        hoja.cell(row=fila, column=2, value=reserva['cliente']).alignment = centrado
        hoja.cell(row=fila, column=3, value=reserva['sala']).alignment = centrado
        hoja.cell(row=fila, column=4, value=reserva['cupo']).alignment = centrado
        hoja.cell(row=fila, column=5, value=reserva['nombre_evento']).alignment = centrado
        hoja.cell(row=fila, column=6, value=reserva['turno']).alignment = centrado
    
    for columna in hoja.columns:
        letra = columna[0].column_letter 
        longitudes = []
        for celda in columna:
            if celda.row == 1:
                continue
            longitudes.append(len(str(celda.value)))
        hoja.column_dimensions[letra].width = max(longitudes) + 2

    hoja.merge_cells("A1:F1")

    try:
        libro.save(f"Reporte-{formato(fecha)}.xlsx")
    except PermissionError:
        print("\nEl archivo está abierto, no se puede sobrescribir.")
    else:
        titulo("Reporte EXCEL generado exitosamente.")


def agregar_reservacion():
    if not hay_registros("clientes") and not hay_registros("salas"):
        print("Aún no hay clientes ni salas agregados para hacer una reservación.")
        return
    if not hay_registros("clientes"):
        print("Aún no hay clientes para hacer una reservación")
        return
    if not hay_registros("salas"):
        print("Aún no hay salas para hacer una reservación")
        return
    while True:
        obtener_clientes()
        id_cliente = entrada("Ingrese la clave del cliente para la reservación", tipo=int)
        if id_cliente is None:
            return
        cliente = existe_cliente(id_cliente)
        if not cliente:
            print("\nNo existe ese cliente, favor de elegir el correcto")
            continue
        break
    while True:
        fecha_reservacion = entrada("Ingrese la fecha de reservación (MM-DD-YYYY) (al menos 2 días posterior a hoy)", tipo=dt.date)
        if fecha_reservacion is None:
            return
        if fecha_reservacion < FECHA_MINIMA:
            print("Esa fecha no cumple con los requisitos (2 días posteriores a hoy)\n")
            continue
        fecha_reservacion = validacion_domingo(fecha_reservacion)
        if fecha_reservacion is None:
            return
        if fecha_reservacion == "N":
            continue
        break   
    while True:
        obtener_salas(fecha_reservacion)
        id_sala = entrada("Ingrese la clave de la sala que desea usar", tipo=int)
        if id_sala is None:
            return
        sala = existe_sala(id_sala)
        if not sala:
            print("\nSala no encontrada, favor de elegir la correcta.")
            continue
        libres = [turno for turno in TURNOS if turno not in turnos_ocupados(id_sala, formato_iso(fecha_reservacion))]
        if not libres:
            print("Esa sala no tiene turnos libres ese día. Elige otra sala.")
            continue
        break
    while True:
        turno_elegido = entrada("Ingrese el turno para el evento (M,V,N)")
        if turno_elegido is None:
            return
        if turno_elegido not in ("M", "V", "N", "Matutino", "Vespertino", "Nocturno"):
            print("No existe ese turno, favor de elegir el correcto (M,V,N)\n")
            continue
        if turno_elegido == "M":
            turno_elegido = "Matutino"
        if turno_elegido == "V":
            turno_elegido = "Vespertino"
        if turno_elegido == "N":
            turno_elegido = "Nocturno"

        ocupados = turnos_ocupados(id_sala, formato_iso(fecha_reservacion))
        if turno_elegido in ocupados:
            print(f"El turno {turno_elegido} está ocupado en la sala {id_sala} con fecha: {formato(fecha_reservacion)}")
            continue
        agregar_reservacion_turno(id_cliente, id_sala, fecha_reservacion, turno_elegido)
        return
    
def editar_nombre_reservacion():
    if not hay_registros("reservas"):
        print("No existen reservaciones aún para editar.")
        return
    while True:
        fecha_inicio = entrada("Ingrese la fecha de inicio para la búsqueda (MM-DD-YYYY)", tipo=dt.date)
        if fecha_inicio is None:
            return
        fecha_fin = entrada("Ingrese la fecha fin de la búsqueda (MM-DD-YYYY)", tipo=dt.date)
        if fecha_fin is None:
            return
        if fecha_inicio > fecha_fin:
            print("La fecha inicial no puede ser mayor a la fecha final.")
            continue
        break
    inicio = formato_iso(fecha_inicio)
    fin = formato_iso(fecha_fin)
    reservas = reservas_en_rango(inicio, fin)
    if not reservas:
        print("\nNo hay reservaciones en ese rango. Volviendo al menú...")
        return
    mostrar_reservacion_fechas(fecha_inicio, fecha_fin, reservas)
    while True:
        editar = entrada("Ingrese el folio del evento para editar su nombre", tipo=int)
        if editar is None:
            return
        folio = existe_folio(editar, inicio, fin)
        if not folio:
            print("\nNo existe ese folio, favor de elegir el correcto")
            continue
        nuevo = entrada("Ingrese el nuevo nombre")
        if nuevo is None:
            return
        actualizar_nombre_evento(editar, nuevo)
        print("El nombre del evento ha sido cambiado exitosamente.")
        return
    
def consultar_reservaciones():
    if not hay_registros("reservas"):
        print("Aún no hay reservaciones para ninguna fecha")
        return
    while True:
        seleccion_fecha = entrada("Ingrese la fecha para ver las reservaciones (MM-DD-YYYY)", tipo=dt.date, hoy="S")
        if seleccion_fecha is None:
            return
        fecha = formato_iso(seleccion_fecha)
        reservas = reservas_por_fecha(fecha)
        if not reservas:
            print("No existe ninguna reservación para esa fecha, intentando de nuevo...\n")
            continue
        titulo2(f"Reservaciones para el día {formato(seleccion_fecha)}")
        print(f"{'Folio':<8}{'Cliente':<19}{'Sala':<20}{'Nombre del evento':<20}{'Turno':<10}")
        sep(90)
        for reserva in reservas:
            print(f"{reserva['folio']:<8}{reserva['cliente']:<19}{reserva['sala']:<20}{reserva['nombre_evento']:<20}{reserva['turno']:<10}")
        sep(90)
        while True:
            exportar = entrada("¿Desea exportar el reporte? (S/N)")
            if exportar is None:
                return
            if exportar not in ("SN"):
                print("No existe esa opción, favor de elegir la correcta (S/N)\n")
                continue
            if exportar == "N":
                print("Volviendo a buscar...\n")
                break
            while True:
                titulo("Exportación de reporte")
                print("1. JSON")
                print("2. CSV")
                print("3. EXCEL")
                opcion = entrada("¿En qué tipo de archivo desea guardar el reporte?", tipo=int)
                if opcion is None:
                    return
                if opcion == 1:
                    exportar_json(reservas,seleccion_fecha)
                    titulo("Reporte JSON generado exitosamente")
                    return
                if opcion == 2:
                    exportar_csv(reservas,seleccion_fecha)
                    titulo("Reporte CSV generado exitosamente")
                    return
                if opcion == 3:
                    exportar_excel(reservas,seleccion_fecha)
                    return
                else:
                    print("\nNo válido, elija entre 1-3.")
                    continue

def agregar_cliente():
    nombre = entrada("Ingrese el nombre del cliente")
    if nombre is None:
        return
    apellido = entrada("Ingrese el apellido del cliente")
    if apellido is None:
        return
    valores = (nombre, apellido)
    id_cliente = insertar_cliente(valores)
    titulo("Cliente agregado exitosamente")
    print(f"Nombre: {nombre} Apellido: {apellido}")
    print(f"Clave: {id_cliente}")
    sep(70)
    salto()
    return

def agregar_sala():
    nombre_sala = entrada("Ingrese el nombre de la sala")
    if nombre_sala is None:
        return
    while True:
        cupo_sala = entrada("Ingrese el cupo de la sala (mayor que 0)", tipo=int)
        if cupo_sala is None:
            return 
        if cupo_sala <1:
            print("La sala debe ser mayor que 0")
            continue
        valores = (nombre_sala, cupo_sala)
        id_sala = insertar_sala(valores)
        titulo("Sala agregada exitosamente")
        print(f"Nombre de la Sala:  {nombre_sala} Cupo: {cupo_sala}")
        print(f"ID: {id_sala}")
        sep(70)
        salto()
        return
    
def salir():
    while True:
        fin = input("¿Desea salir del programa? (S/N): ").title().strip()
        if fin not in ("S","N") or not fin:
            print("No es válido, favor de usar (S/N)\n")
            continue
        if fin == "N":
            print("Regresando al menú...\n")
            return
        if fin == "S":
            titulo("Programa terminado...")
            sys.exit()

def inicio():
    clientes_count = db.execute_select("SELECT COUNT(*) FROM clientes")[0][0]
    salas_count = db.execute_select("SELECT COUNT(*) FROM salas")[0][0]
    reservas_count = db.execute_select("SELECT COUNT(*) FROM reservas")[0][0]

    if clientes_count == 0 and salas_count == 0 and reservas_count == 0:
        titulo("Estado inicial del sistema")
        print("No se encontró un estado previo en la base de datos.")
        print("Se inicia con un estado vacío (sin clientes, salas ni reservaciones).")
        sep(70)
    else:
        titulo("Estado cargado correctamente")
        print(f"Clientes registrados: {clientes_count}")
        print(f"Salas registradas: {salas_count}")
        print(f"Reservaciones existentes: {reservas_count}")
        sep(70)
        
        
def menu():
    try: 
        while True:
            titulo("MENÚ PRINCIPAL")
            print("1. Registrar la reservación de una sala.")
            print("2. Editar el nombre del evento de una reservación ya hecha.")
            print("3. Consultar las reservaciones existentes para una fecha específica. ")
            print("4. Registrar a un nuevo cliente")
            print("5. Registrar una sala")
            print("6. Salir")
            try: 
                respuesta = int(input("Ingrese la opción: "))
            except ValueError:
                print("Error: Escribir solo números del 1-6")
            else:
                if respuesta == 1:
                    agregar_reservacion()
                elif respuesta == 2:
                    editar_nombre_reservacion()
                elif respuesta == 3:
                    consultar_reservaciones()
                elif respuesta == 4:
                    agregar_cliente()
                elif respuesta == 5:
                    agregar_sala()
                elif respuesta == 6:
                    salir()
                else:
                    print("Fuera de rango, elegir solo entre 1-6")
    except KeyboardInterrupt:
        titulo("Cierre forzado...")


if __name__ == "__main__":
    conexion()
    inicio()
    menu()